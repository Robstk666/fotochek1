from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes
import requests
import openpyxl
from io import BytesIO
from datetime import datetime
import re

# Укажите токен Telegram-бота и OAuth-токен Яндекс.Диска
TELEGRAM_TOKEN = '6454785033:AAExJyrYHsOyS2LJLjAcnaKIY2_O5vsAYwY'
YANDEX_TOKEN = 'y0_AgAAAAAfhlAuAAxsegAAAAEQoT8yAAA2fjIZkRpDWrJFD2pxkYSVuElzZw'

# Доступные стройки
CONSTRUCTION_SITES = ['Стройка 1', 'Стройка 2', 'Стройка 3']


# Функция для загрузки файла на Яндекс.Диск
def upload_to_yandex_disk(file_data: BytesIO, filename: str) -> str:
    headers = {'Authorization': f'OAuth {YANDEX_TOKEN}'}
    upload_url = "https://cloud-api.yandex.net/v1/disk/resources/upload"
    params = {"path": f"disk:/Загрузки/{filename}", "overwrite": "true"}

    # Получаем ссылку для загрузки
    response = requests.get(upload_url, headers=headers, params=params)
    if response.status_code == 200:
        href = response.json().get("href")
        # Загружаем файл на Яндекс.Диск
        upload_response = requests.put(href, files={"file": file_data})
        if upload_response.status_code == 201:
            print("Файл успешно загружен на Яндекс.Диск.")
            return f"https://disk.yandex.ru/client/disk/{filename}"
        else:
            print(f"Ошибка при загрузке файла на Яндекс.Диск: {upload_response.text}")
            return "Ошибка при загрузке файла"
    else:
        print(f"Ошибка при получении ссылки для загрузки на Яндекс.Диск: {response.text}")
        return "Ошибка при получении ссылки"


# Функция для сохранения данных в Excel на разных листах
def save_to_excel(construction_site, date, amount, link):
    excel_file = 'data.xlsx'

    # Проверяем, существует ли файл
    try:
        workbook = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Проверяем, существует ли лист для стройки
    if construction_site not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=construction_site)
        sheet.append(['Дата', 'Сумма', 'Ссылка на файл'])
    else:
        sheet = workbook[construction_site]

    # Добавляем данные
    sheet.append([date, amount, link])
    workbook.save(excel_file)
    print(f"Данные успешно сохранены в Excel на листе {construction_site}.")


# Обработчик команды /1 (вместо /start)
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [[InlineKeyboardButton(site, callback_data=site) for site in CONSTRUCTION_SITES]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text('Привет! Выберите стройку:', reply_markup=reply_markup)


# Обработчик выбора стройки
async def choose_construction_site(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data['construction_site'] = query.data
    await query.edit_message_text(
        text=f"Вы выбрали: {query.data}. Теперь введите сумму чека в рублях (целое число, например, 1234).")


# Обработчик текстовых сообщений
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if 'construction_site' not in context.user_data:
        await update.message.reply_text("Сначала выберите стройку с помощью команды /1.")
        return

    text = update.message.text

    # Проверяем формат ввода суммы (целое число)
    if not re.match(r'^\d+$', text):
        await update.message.reply_text("Пожалуйста, введите сумму в рублях как целое число (например, 1234).")
        return

    context.user_data['amount'] = text
    await update.message.reply_text(f"Сумма {text} рублей сохранена. Теперь отправьте фото чека.")


# Обработчик фото
async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if 'amount' not in context.user_data:
        await update.message.reply_text("Сначала отправьте сумму чека.")
        return

    amount = context.user_data['amount']
    construction_site = context.user_data['construction_site']
    photo = update.message.photo[-1]
    file = await photo.get_file()
    file_data = BytesIO()
    await file.download_to_memory(file_data)
    file_data.seek(0)

    filename = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + ".jpg"
    await update.message.reply_text("Начинаем загрузку фото на Яндекс.Диск...")

    link = upload_to_yandex_disk(file_data, filename)
    if "Ошибка" not in link:
        await update.message.reply_text(f"Фото успешно загружено. Ссылка: {link}")
        save_to_excel(construction_site, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), amount, link)
        await update.message.reply_text(f"Спасибо! Ваш чек и сумма {amount} рублей успешно сохранены в системе.")
    else:
        await update.message.reply_text(link)


# Обработчик команды /6 для отправки Excel-файла (замена /getfile)
async def getfile(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat_id
    excel_file = 'data.xlsx'

    try:
        with open(excel_file, 'rb') as file:
            await context.bot.send_document(chat_id=chat_id, document=file, filename=excel_file)
        await update.message.reply_text("Вот ваш файл Excel со всеми стройками.")
    except FileNotFoundError:
        await update.message.reply_text("Файл Excel не найден. Сначала добавьте данные.")


# Обработчик команды /welcome для приветствия
async def welcome(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Добро пожаловать! Наберите /1 для начала работы с ботом.")


# Основная функция
def main():
    application = ApplicationBuilder().token(TELEGRAM_TOKEN).build()

    # Обработчики команд и сообщений
    application.add_handler(CommandHandler("1", start))
    application.add_handler(CommandHandler("6", getfile))  # Изменение здесь
    application.add_handler(CommandHandler("welcome", welcome))
    application.add_handler(CallbackQueryHandler(choose_construction_site))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    application.add_handler(MessageHandler(filters.PHOTO, handle_photo))

    print("Запуск бота...")
    application.run_polling()


if __name__ == '__main__':
    main()